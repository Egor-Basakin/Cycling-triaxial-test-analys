import tkinter as tk
from tkinter import filedialog
import pandas as pd
import numpy as np
from scipy.signal import find_peaks
from scipy.optimize import curve_fit
from tkinter import ttk
from tkinter import simpledialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import csv
import os
import shutil
import threading
import time
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series



df = None  # Объявление переменной df в глобальной области видимости
csv_files = []

# Функции для обработки событий


# Функция для получения списка файлов CSV в директории
def list_csv_files(directory):
    return [file for file in os.listdir(directory) if file.endswith('.csv')]

# Функция для обновления списка файлов
def update_file_list():
    global csv_files
    while True:
        csv_files = list_csv_files(directory_path)  # Обновляем список файлов
        file_listbox.delete(0, tk.END)  # Очищаем список в GUI
        for file in csv_files:
            file_listbox.insert(tk.END, file)  # Добавляем файлы в список в GUI
        # Приостанавливаем выполнение на заданный интервал времени (например, 1 секунда)
        time.sleep(5)


def on_file_select(event):
    global df
    selected_file = file_listbox.get(file_listbox.curselection())
    full_path = os.path.join(directory_path, selected_file)  # Формирование полного пути к файлу
    print(f"Выбран файл: {full_path}")
     # Загрузка данных из файла
    try:
        df = pd.read_csv(full_path)  # или измените на соответствующий формат файла, если не CSV
        max_pick()  # Вызов функции max_pick после успешной загрузки файла
    except Exception as e:
        print("Ошибка при загрузке файла:", e)


def import_file_geo():
    global df
    file_path = filedialog.askopenfilename()
    # Загрузить данные (возможно используется кодтровака, и нужно указать другой разделитель между данными df = pd.read_csv('sinusoida.csv', encoding='utf-8', quotechar='"')
    print("Загрузка данных...")
    source_file = file_path

    data = pd.read_csv(source_file, sep='\t', decimal=',')
    data = data[['Time', 'ChVerticalDeformation_mm']]
    data = data.dropna(subset=['ChVerticalDeformation_mm'])
        # Фильтрация строк с значениями, имеющими три знака после запятой
    data = data[data['Time'].astype(str).str.contains(r'\.\d{3}$')]
    df = data

    print(df)

    print("Данные загружены.")

def import_file_gds():
    global df
    file_path = filedialog.askopenfilename()
    # Загрузить данные (возможно используется кодтровака, и нужно указать другой разделитель между данными df = pd.read_csv('sinusoida.csv', encoding='utf-8', quotechar='"')
    print("Загрузка данных...")
    source_file = file_path
    # Желаемое новое расширение (например, .csv)
    new_extension = '.csv'

    # Изменяем расширение файла
    base_name, _ = os.path.splitext(source_file)
    new_file = base_name + new_extension

    # Копируем исходный файл с новым именем и расширением
    shutil.copy(source_file, new_file)
    # Теперь у вас есть файл с новым расширением (new_file)

    try:
        with open(new_file, 'r') as file:
            lines = file.readlines()
    except FileNotFoundError:
        print(f"Файл {new_file} не найден.")
        exit(1)

    # Удаление последней строки (если она есть)
    if len(lines) > 1:
        lines = lines[:-1]

    # Запись обновленных строк обратно в файл CSV
    with open(new_file, 'w') as file:
        file.writelines(lines)


    # Например, для чтения нового файла:
    with open(new_file, 'r') as file:
        data = pd.read_csv(new_file, skiprows=28, skipfooter=3)
        data = data[['Time since start of test (s)', 'Axial Displacement (mm)']] #загрузил толь эти два столбца
        data = data[data['Time since start of test (s)'] % 1 != 0]    #отсортирвоал только дробные
        data = data.rename(columns={'Time since start of test (s)': 'Time', 'Axial Displacement (mm)': 'ChVerticalDeformation_mm'})  #переименовал столбцы
    df = data

    print("Данные загружены.")

def save1():
    global df # Объявление, что вы используете глобальную переменную df
    if df is not None:
        save_file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("csv_files", "*.csv"), ("All files", "*.*")])
        if save_file_path:
            df.to_csv(save_file_path, index=False)
            print("Данные сохранены в файл:", save_file_path)
            update_file_list()
        else:
            print("Сохранение отменено")
    else:
        print("Нет данных для сохранения")

def import_and_save_geo():
    import_file_geo()
    save1()

def import_and_save_gds():
    import_file_gds()
    save1()


def max_pick():
    global peak_rows
    if df is not None:
        # Предполагаем, что данные о синусоиде находятся в столбце 'Axial Displacement (mm)'
        sinusoid_column = 'ChVerticalDeformation_mm'
        if sinusoid_column in df.columns:
            print(f"Анализ столбца: {sinusoid_column}")

            # Найти индексы пиков
            print("Поиск пиков...")
            peaks, _ = find_peaks(df[sinusoid_column])
            print(f"Найдено {len(peaks)} пиков.")

            # Выделить строки с пиками
            print("Выделение строк с максимальными пиками...")
            peak_rows = df.iloc[peaks]
            # Дополнительный код для работы с peak_rows
        else:
            print(f"Столбец {sinusoid_column} не найден в файле.")
    else:
        print("Данные не загружены. Пожалуйста, выберите файл.")

def on_file_select_and_seash_peak():
    global df
    global peak_rows
    on_file_select()
    max_pick()

def filtr_1():
    global new_data
    global peak_rows
    try:
        step = int(step_entry.get()) #получаем шаг из виджета Entry
        if step > 0:
            print("Выделение каждого " + str(step) + " пика...")
            # Фильтрация данных, чтобы учитывать только значения с дробной частью
            data_filtered = peak_rows[peak_rows['Time'] % 1 != 0]
            # Выбор каждого пвторого пика ( [::2])
            every_fifth_peak = data_filtered.iloc[::step]
        else:
            raise ValueError("Шаг должен быть больше 0")
    except ValueError as e:
        tk.messagebox.showerror("Ошибка ввода", "Введите корректный шаг (целое число больше 0)")

    # Отобразить результаты (все пики)
    #print("Готово. Вот строки с максимальными пиками:")
    #print(peak_rows)

    # Отобразить результаты (каждый 2 пик)
    print("Готово. Вот каждая " + str(step) +" строка с максимальным пиком:")
    print(every_fifth_peak)

#def creat_column():
        # Фиксация значения первой строки ('Time since start of test (s)')
    first_row_Time_since = every_fifth_peak.iloc[0]['Time']

    # Вычисление разницы и создание нового столбца времени (Time)
    every_fifth_peak['Cycle'] = (every_fifth_peak['Time'] - first_row_Time_since) * 1
    #создание столбца логарифма времени ln_N
    every_fifth_peak['ln_N'] = np.log(every_fifth_peak['Cycle'])

    # Фиксация значения первой строки ('Axial Displacement (mm)')
    first_row_axial = every_fifth_peak.iloc[0]['ChVerticalDeformation_mm']
    # Вычисление высоты образца (height)
    height = 100 - first_row_axial
    # Вычисление и создание нового столбца осевой деформации (axial_deformation)
    every_fifth_peak['axial_deformation'] = ((every_fifth_peak['ChVerticalDeformation_mm'] - first_row_axial) / height) 

    #создаю новую перменную, содержащую добавленные столбцы new_data
    new_data = every_fifth_peak

    # Предложить пользователю сохранить результаты в файл Excel
    #save_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
    #                                              filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])


def plt_1():
    global new_data
            # Извлеките значения x и y из DataFrame (new_data)
    x = new_data['ln_N']
    y = new_data['axial_deformation']

     # Постройте график
    plt.figure()
    plt.plot(x, y)

    plt.legend()

    # Добавьте заголовок и подписи к осям
    plt.title('Пример графика')
    plt.xlabel('Ось X')
    plt.ylabel('Ось Y')
    # Добавление сетки
    plt.grid(True)
    # Отобразите график
    plt.show()

def approx():
    global every_fifth_peak
    global new_data
    global equation_text_1

    x = new_data['ln_N']
    y = new_data['axial_deformation']
     # Запрашиваем координаты точек для аппроксимации
    print("Введите координаты двух точек для аппроксимации.")
    
    x1 = float(x1_entry.get())
    x2 = float(x2_entry.get())
    # Выборка данных для аппроксимации
    mask = (x >= x1) & (x <= x2)
    x_approx = x[mask]
    y_approx = y[mask]
          
    # Определение логарифмической функции для аппроксимации
    #def log_func(N, a, b):
    #    return a * np.log(N) + (b)
    # Определение линейной функции для аппроксимации
    def linear_func(N, a, b):
        return a * N + (b)
        
    # Аппроксимация данных
    popt, _ = curve_fit(linear_func, x_approx, y_approx)

#def plt_2():
   
    # Постройте график с логарифмической осью
    plt.plot(x, y, color='black')
    plt.xlabel('ln(N), циклов')
    plt.ylabel('Осевая деформация, \u03B5 1 доли ед.')

    # Отображение аппроксимирующей линии
    plt.plot(x_approx, linear_func(x_approx, *popt), label='Аппроксимирующая кривая', color='red', linewidth=2, zorder=2)
    plt.legend()


    # Показ уравнения на графике
    equation = f"\u03B5 = {popt[0]:.6f} * ln(N) {popt[1]:.6f}"
    plt.text(0.05, 0.80, equation, transform=plt.gca().transAxes, fontsize=12, color='black', verticalalignment='top')



    # Установка основных делений сетки на 10^0, 10^1, 10^2 и так далее
    plt.grid()

    # Показать график
    plt.show()

    # Добавляем текстовую часть уравнения линии тренда
    equation_text_1 = "\u03B5 = {popt[0]:.6f} * ln(N) {popt[1]:.6f}"

      # Создание новых столбцов под аппроксимацию данных
    new_data['x_approx']=x_approx
    new_data['y_approx']=y_approx

    # Создание столбца для пострения ааппроксимирующей линии
    new_data['y_approx_linear'] = linear_func(x_approx, popt[0], popt[1])

    

    
def save2():
    global save_file_path, new_data, equation_text_1
    
    # Предложить пользователю сохранить результаты в файл Excel
    save_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                  filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if save_file_path:
    # Сохранение данных в Excel
        new_data.to_excel(save_file_path, index=False)
 

#def plot_excel():
 #   global new_data
  #  global save_file_path
    # Загружаем существующий Excel-файл
        wb = load_workbook(save_file_path)
        ws = wb.active

       # Создаем точечную диаграмму
        chart = ScatterChart()
        chart.title = "График"
        chart.style = 13  # Гладкие линии

        # Данные для первой серии (6-й столбец для X, 8-й столбец для Y)
        x_values = Reference(ws, min_col=6, min_row=3, max_row=ws.max_row)
        y_values = Reference(ws, min_col=8, min_row=3, max_row=ws.max_row)
        series_1 = Series(y_values, xvalues=x_values, title_from_data=True)
        series_1.smooth = True  # Гладкая кривая для первой серии
        series_1.graphicalProperties.line.solidFill = "990000"  # Красный цвет линии
        chart.series.append(series_1)

        # Данные для второй серии (4-й столбец для X, 5-й столбец для Y)
        x_values_2 = Reference(ws, min_col=4, min_row=3, max_row=ws.max_row)
        y_values_2 = Reference(ws, min_col=5, min_row=3, max_row=ws.max_row)
        series_2 = Series(y_values_2, xvalues=x_values_2, title_from_data=True)
        series_2.smooth = True  # Гладкая кривая для второй серии
        series_2.graphicalProperties.line.width = 20000  # Ширина линии (значение по умолчанию - 100000)
        series_2.graphicalProperties.line.solidFill = "000000"  # Черный цвет линии
        chart.series.append(series_2)



        # Добавляем график в лист
        ws.add_chart(chart, "K2")  # Расположение графика

        # Сохраняем изменения в файле
        wb.save(save_file_path)
        print("Данные сохранены в файл:", save_file_path)
    else:
        print("Сохранение отменено")






# Создание главного окна
root = tk.Tk()
root.title("Обработчик v.0.2")

# Рамка для списка файлов
frame_files = ttk.Frame(root)
frame_files.pack(side=tk.LEFT, fill=tk.Y, expand=False)

# Список файлов
file_listbox = tk.Listbox(frame_files)
file_listbox.pack(side="left", fill="both", expand=True)


directory_path= "C:/Users/Егор/Desktop/python/file"

update_thread = threading.Thread(target=update_file_list)
update_thread.daemon = True  # Поток завершится при завершении основной программы
update_thread.start()

# Привязка функции к событию выбора файла
file_listbox.bind('<<ListboxSelect>>', on_file_select)

# Кнопка для функции max_pick
#max_pick_button = tk.Button(root, text="Найти максимальные пики", command=max_pick)
#max_pick_button.pack()

# Добавление кнопки для построения графика
plot_button = tk.Button(root, text="Построить график-1", command=plt_1)
plot_button.pack()


# Виджет для ввода шага
step_label = tk.Label(root, text="Введите шаг:")
step_label.pack()
step_entry = tk.Entry(root)
step_entry.pack()

# Кнопка для запуска фильтрации
filtr_button = tk.Button(root, text="Применить фильтр", command=filtr_1)
filtr_button.pack()

# Виджеты для ввода X1 и X2
x1_label = tk.Label(root, text="Введите X первой точки:")
x1_label.pack()
x1_entry = tk.Entry(root)
x1_entry.pack()

x2_label = tk.Label(root, text="Введите X второй точки:")
x2_label.pack()
x2_entry = tk.Entry(root)
x2_entry.pack()

# Кнопка для аппроксимации
approx_button = tk.Button(root, text="Выполнить аппроксимацию", command=approx)
approx_button.pack()

# Кнопка для сохранения файла Эксель
save_to_exel_button = tk.Button(root, text="Сохранить в формате Excel", command=save2)
save_to_exel_button.pack()


# Создание меню
menu_bar = tk.Menu(root)
root.config(menu=menu_bar)

file_menu = tk.Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="Файл", menu=file_menu)
file_menu.add_command(label="Импортирвать Геотек", command=import_and_save_geo)
file_menu.add_command(label="Импортировать GDS", command=import_and_save_gds)
file_menu.add_separator()
file_menu.add_command(label="Выход", command=root.quit)

# Создание текстового поля
text = tk.Text(root)
text.pack()

# Запуск главного цикла
root.mainloop()
